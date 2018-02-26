from selenium import webdriver
from openpyxl import Workbook


wb = Workbook()
ws1 = wb.active
ws1.title = 'first_ Sheet'


#try ~ except는 무시해도 되는 구문입니다.
#그냥 드라이버를 불러온다고 생각하면 됩니다.
try:
    driver = webdriver.Chrome('./driver/chromedriver.exe')
except:
    driver = webdriver.Chrome('./driver/chromedriver')


#yes24 2017년 12월 월별 베스트 페이지로 이동
driver.get('http://www.yes24.com/24/category/bestseller?CategoryNumber=001001003&sumgb=09&year=2017&month=12')
#세부 페이지

for i in range(1, 7, 2):
    book = driver.find_element_by_xpath(
        '//*[@id="category_layout"]/tbody/tr[%s]/td[3]/p[1]/a[1]' % i)
    print(book.text)
    s = (i+1)/2
    ws1.cell(row=s,  column=1, value=book.text)
    book.click()
    try:
        page_gram_size = driver.find_element_by_xpath(
            '// *[ @ id = "tblGoodsFairTraderNoti"] / tbody / tr[2] / td')
        print(page_gram_size.text)
        s = (i + 1) / 2
        ws1.cell(row=s,  column=2, value=page_gram_size.text)
    except:
        pass
    try:
        book_introduce = driver.find_element_by_xpath(
            '// *[ @ id = "contents"] / div[4] / p[1]')
        print(book_introduce.text)
        s = (i + 1) / 2
        ws1.cell(row=s,  column=3, value=book_introduce.text)
    except:
        pass
    try:
        intro = driver.find_element_by_xpath(
            '//*[@id="contents_constitution_text0"]/span[1]')
        print(intro.text)
        s = (i + 1) / 2
        ws1.cell(row=s,  column=4, value=intro.text)
    except:
        pass
    driver.back()
wb.save("c:/excel/test11.xlsx")
